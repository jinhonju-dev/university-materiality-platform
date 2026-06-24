import csv
from io import BytesIO, StringIO

from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from .analytics import build_analytics
from .models import (
    ConcernSurveyResponse,
    ConcernSurveyScore,
    ExpertAssessmentResponse,
    ExpertAssessmentScore,
    InvitationCode,
    StakeholderGroup,
    SurveyCampaign,
    SurveyResponse,
    Topic,
    TopicScore,
    User,
)


LEGACY_RAW_HEADERS = [
    "response_id",
    "submitted_at",
    "respondent_email",
    "respondent_name",
    "invitation_code_id",
    "stakeholder_group",
    "stakeholder_weight",
    "topic_code",
    "topic_name_zh",
    "category",
    "organization_score",
    "impact_score",
    "financial_score",
    "open_answer",
]


CONCERN_HEADERS = [
    "response_id",
    "submitted_at",
    "stakeholder_group",
    "stakeholder_weight",
    "topic_code",
    "topic_name_zh",
    "category",
    "concern_score",
    "open_answer",
]


EXPERT_HEADERS = [
    "response_id",
    "submitted_at",
    "invitation_code_id",
    "evaluator_role",
    "stakeholder_group",
    "stakeholder_weight",
    "topic_code",
    "topic_name_zh",
    "category",
    "positive_likelihood_score",
    "positive_impact_magnitude_score",
    "negative_likelihood_score",
    "negative_impact_magnitude_score",
    "impact_score",
    "enrollment_revenue_score",
    "reputation_score",
    "operating_cost_score",
    "funding_score",
    "legal_responsibility_score",
    "financial_likelihood_score",
    "financial_score",
    "open_answer",
]


def append_sheet(workbook: Workbook, title: str, headers: list[str], rows: list[list]) -> None:
    sheet = workbook.create_sheet(title=title[:31])
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    for column in sheet.columns:
        length = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max(length + 2, 12), 42)


def workbook_bytes(workbook: Workbook) -> BytesIO:
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def legacy_raw_rows(db: Session, campaign: SurveyCampaign, anonymized: bool) -> list[list]:
    rows = db.execute(
        select(SurveyResponse, TopicScore, Topic, StakeholderGroup, User)
        .join(TopicScore, TopicScore.response_id == SurveyResponse.id)
        .join(Topic, Topic.id == TopicScore.topic_id)
        .join(StakeholderGroup, StakeholderGroup.id == SurveyResponse.stakeholder_group_id)
        .outerjoin(User, User.id == SurveyResponse.respondent_id)
        .where(SurveyResponse.campaign_id == campaign.id)
        .order_by(SurveyResponse.id, Topic.sort_order)
    ).all()
    output = []
    for response, score, topic, group, user in rows:
        output.append(
            [
                response.id,
                response.submitted_at.isoformat(),
                "" if anonymized else (user.email if user else ""),
                "" if anonymized else (user.name if user else ""),
                response.invitation_code_id or "",
                group.name,
                group.weight,
                topic.topic_code or topic.code,
                topic.name_zh,
                topic.category,
                score.organization_score,
                score.impact_score,
                score.financial_score,
                response.open_answer or "",
            ]
        )
    return output


def concern_rows(db: Session, campaign: SurveyCampaign) -> list[list]:
    rows = db.execute(
        select(ConcernSurveyResponse, ConcernSurveyScore, Topic, StakeholderGroup)
        .join(ConcernSurveyScore, ConcernSurveyScore.response_id == ConcernSurveyResponse.id)
        .join(Topic, Topic.id == ConcernSurveyScore.topic_id)
        .join(StakeholderGroup, StakeholderGroup.id == ConcernSurveyResponse.stakeholder_group_id)
        .where(ConcernSurveyResponse.campaign_id == campaign.id)
        .order_by(ConcernSurveyResponse.id, Topic.sort_order)
    ).all()
    return [
        [
            response.id,
            response.submitted_at.isoformat(),
            group.name,
            group.weight,
            topic.topic_code or topic.code,
            topic.name_zh,
            topic.category,
            score.concern_score,
            response.open_answer or "",
        ]
        for response, score, topic, group in rows
    ]


def expert_rows(db: Session, campaign: SurveyCampaign, anonymized: bool = False) -> list[list]:
    rows = db.execute(
        select(ExpertAssessmentResponse, ExpertAssessmentScore, Topic, StakeholderGroup, InvitationCode)
        .join(ExpertAssessmentScore, ExpertAssessmentScore.response_id == ExpertAssessmentResponse.id)
        .join(Topic, Topic.id == ExpertAssessmentScore.topic_id)
        .join(StakeholderGroup, StakeholderGroup.id == ExpertAssessmentResponse.stakeholder_group_id)
        .join(InvitationCode, InvitationCode.id == ExpertAssessmentResponse.invitation_code_id)
        .where(ExpertAssessmentResponse.campaign_id == campaign.id)
        .order_by(ExpertAssessmentResponse.id, Topic.sort_order)
    ).all()
    return [
        [
            response.id,
            response.submitted_at.isoformat(),
            "" if anonymized else response.invitation_code_id,
            invitation.evaluator_role or "",
            group.name,
            group.weight,
            topic.topic_code or topic.code,
            topic.name_zh,
            topic.category,
            (score.positive_likelihood_score if score.positive_likelihood_score is not None else score.impact_likelihood_score) or "",
            (score.positive_impact_magnitude_score if score.positive_impact_magnitude_score is not None else score.positive_impact_score) or "",
            (score.negative_likelihood_score if score.negative_likelihood_score is not None else score.impact_likelihood_score) or "",
            (score.negative_impact_magnitude_score if score.negative_impact_magnitude_score is not None else score.negative_impact_score) or "",
            score.impact_score,
            (score.enrollment_revenue_score if score.enrollment_revenue_score is not None else score.admissions_revenue_score) or "",
            score.reputation_score or "",
            score.operating_cost_score or "",
            score.funding_score or "",
            (score.legal_responsibility_score if score.legal_responsibility_score is not None else score.legal_liability_score) or "",
            score.financial_likelihood_score or "",
            score.financial_score,
            response.open_answer or "",
        ]
        for response, score, topic, group, invitation in rows
    ]


def create_csv_export(db: Session, campaign: SurveyCampaign, anonymized: bool = False) -> BytesIO:
    text = StringIO()
    writer = csv.writer(text)
    writer.writerow(LEGACY_RAW_HEADERS)
    writer.writerows(legacy_raw_rows(db, campaign, anonymized))
    output = BytesIO()
    output.write(text.getvalue().encode("utf-8-sig"))
    output.seek(0)
    return output


def append_analytics_sheets(workbook: Workbook, db: Session, campaign: SurveyCampaign) -> None:
    analytics = build_analytics(db, campaign)
    append_sheet(
        workbook,
        "topic_averages",
        ["code", "name", "category", "impact", "financial", "weighted_impact", "weighted_financial", "response_count", "quadrant"],
        [
            [
                item["code"],
                item["name"],
                item["category"],
                item["impact"],
                item["financial"],
                item["weighted_impact"],
                item["weighted_financial"],
                item["response_count"],
                item["quadrant"],
            ]
            for item in analytics["topics"]
        ],
    )
    append_sheet(
        workbook,
        "stakeholder_samples",
        ["stakeholder_group", "sample_count", "weight"],
        [[item["name"], item["count"], item["weight"]] for item in analytics["stakeholders"]],
    )
    append_sheet(
        workbook,
        "keyword_counts",
        ["keyword", "count"],
        [[item["keyword"], item["count"]] for item in analytics["keywords"]],
    )


def create_excel_export(db: Session, campaign: SurveyCampaign) -> BytesIO:
    workbook = Workbook()
    workbook.remove(workbook.active)
    analytics = build_analytics(db, campaign)
    concern_campaign = analytics.get("concern_campaign")
    expert_campaign = analytics.get("expert_campaign")
    append_sheet(workbook, "concern_raw", CONCERN_HEADERS, concern_rows(db, concern_campaign) if concern_campaign else [])
    append_sheet(workbook, "concern_anonymized", CONCERN_HEADERS, concern_rows(db, concern_campaign) if concern_campaign else [])
    append_sheet(workbook, "expert_raw", EXPERT_HEADERS, expert_rows(db, expert_campaign, anonymized=False) if expert_campaign else [])
    append_sheet(workbook, "expert_anonymized", EXPERT_HEADERS, expert_rows(db, expert_campaign, anonymized=True) if expert_campaign else [])
    append_sheet(
        workbook,
        "topic_concern_scores",
        ["code", "name", "category", "concern_score", "concern_response_count"],
        [[item["code"], item["name"], item["category"], item["concern_score"], item["concern_response_count"]] for item in analytics["topics"]],
    )
    append_sheet(
        workbook,
        "topic_impact_scores",
        ["code", "name", "category", "impact_materiality_score", "quadrant"],
        [[item["code"], item["name"], item["category"], item["impact_materiality_score"], item["quadrant"]] for item in analytics["topics"]],
    )
    append_sheet(
        workbook,
        "topic_financial_scores",
        ["code", "name", "category", "financial_materiality_score", "quadrant"],
        [[item["code"], item["name"], item["category"], item["financial_materiality_score"], item["quadrant"]] for item in analytics["topics"]],
    )
    append_sheet(
        workbook,
        "stakeholder_segments",
        ["stakeholder_group", "sample_count", "weight"],
        [[item["name"], item["count"], item["weight"]] for item in analytics["stakeholders"]],
    )
    append_sheet(
        workbook,
        "evaluator_role_segments",
        ["evaluator_role", "sample_count"],
        [[item["evaluator_role"], item["count"]] for item in analytics["evaluator_roles"]],
    )
    append_sheet(
        workbook,
        "unknown_ratios",
        ["code", "name", "unknown_ratio_percent"],
        [[item["code"], item["name"], item["unknown_ratio"]] for item in analytics["topics"]],
    )
    append_sheet(
        workbook,
        "final_material_topics",
        ["code", "name", "impact", "financial", "concern", "quadrant", "reason", "manual"],
        [
            [
                item["code"],
                item["name"],
                item["impact_materiality_score"],
                item["financial_materiality_score"],
                item["concern_score"],
                item["quadrant"],
                item["final_topic_reason"] or "",
                item["manually_adjusted"],
            ]
            for item in analytics["final_material_topics"]
        ],
    )
    return workbook_bytes(workbook)


def create_concern_export(db: Session, campaign: SurveyCampaign) -> BytesIO:
    workbook = Workbook()
    workbook.remove(workbook.active)
    append_sheet(workbook, "concern_responses", CONCERN_HEADERS, concern_rows(db, campaign))
    return workbook_bytes(workbook)


def create_expert_export(db: Session, campaign: SurveyCampaign) -> BytesIO:
    workbook = Workbook()
    workbook.remove(workbook.active)
    append_sheet(workbook, "expert_responses", EXPERT_HEADERS, expert_rows(db, campaign, anonymized=False))
    append_sheet(workbook, "expert_anonymized", EXPERT_HEADERS, expert_rows(db, campaign, anonymized=True))
    append_analytics_sheets(workbook, db, campaign)
    return workbook_bytes(workbook)


def create_anonymized_export(db: Session, campaign: SurveyCampaign) -> BytesIO:
    workbook = Workbook()
    workbook.remove(workbook.active)
    append_sheet(workbook, "concern_anonymized", CONCERN_HEADERS, concern_rows(db, campaign))
    append_sheet(workbook, "expert_anonymized", EXPERT_HEADERS, expert_rows(db, campaign, anonymized=True))
    append_sheet(workbook, "legacy_anonymized", LEGACY_RAW_HEADERS, legacy_raw_rows(db, campaign, anonymized=True))
    return workbook_bytes(workbook)
